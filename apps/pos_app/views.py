# pos_app/views.py

from rest_framework import generics, status, views, permissions, filters
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, Count, Q, F, Avg
from django.db import transaction
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import logging

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

from .models import (
    Store, Role, Category, Product,
    Invoice, InvoiceItem, SyncLog, DailySales
)
from .serializers import (
    StoreSerializer, StoreListSerializer, RoleSerializer,
    CategorySerializer, ProductSerializer, ProductListSerializer,
    InvoiceSerializer, InvoiceListSerializer, BulkInvoiceSyncSerializer,
    DashboardStatsSerializer, SalesReportSerializer, ProductReportSerializer,
    SyncLogSerializer, UserProfileSerializer
)
from .permissions import (
    IsOwner, IsOwnerOrReadOnly, IsSameStore,
    IsSalespersonOrOwner, CanCreateInvoice, CanViewReports
)

# Set up logging
logger = logging.getLogger(__name__)



# ============================================
# STORE & ROLE VIEWS
# ============================================

class StoreListView(generics.ListAPIView):
    """
    Public endpoint to list all active stores.
    Used during registration to select a store.
    """
    queryset = Store.objects.filter(is_active=True)
    serializer_class = StoreListSerializer
    permission_classes = []  # Public endpoint
    pagination_class = None  # No pagination for store list


class StoreDetailView(generics.RetrieveAPIView):
    """
    Get authenticated user's store details.
    """
    serializer_class = StoreSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        """Return the store of the authenticated user"""
        return self.request.user.store


class RoleListView(generics.ListAPIView):
    """
    Public endpoint to list all available roles.
    Used during registration to select a role.
    """
    queryset = Role.objects.all().order_by('name')
    serializer_class = RoleSerializer
    permission_classes = []  # Public endpoint
    pagination_class = None  # No pagination for role list


# ============================================
# CATEGORY VIEWS
# ============================================

class CategoryListCreateView(generics.ListCreateAPIView):
    """
    List and create product categories.
    Users can only see and create categories for their store.
    """
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        """Return only active categories from user's store"""
        return Category.objects.filter(
            store=self.request.user.store,
            is_active=True
        ).select_related('parent', 'store')

    def perform_create(self, serializer):
        """Set store automatically from authenticated user"""
        try:
            serializer.save(store=self.request.user.store)
            logger.info(
                f"Category '{serializer.instance.name}' created by user {self.request.user.email}"
            )
        except Exception as e:
            logger.error(f"Error creating category: {str(e)}")
            raise


class CategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Get, update, or soft-delete a category.
    """
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated, IsSameStore]

    def get_queryset(self):
        """Return categories from user's store"""
        return Category.objects.filter(
            store=self.request.user.store
        ).select_related('parent', 'store')

    def perform_destroy(self, instance):
        """Soft delete - mark as inactive instead of deleting"""
        instance.is_active = False
        instance.save()
        logger.info(
            f"Category '{instance.name}' deactivated by user {self.request.user.email}"
        )


# ============================================
# PRODUCT VIEWS
# ============================================

class ProductListCreateView(generics.ListCreateAPIView):
    """
    List and create products.
    Users can only see and create products for their store.
    """
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'is_active']
    search_fields = ['name', 'code', 'barcode', 'description']
    ordering_fields = ['name', 'price', 'stock', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        """Return products from user's store with optimized queries"""
        queryset = Product.objects.filter(
            store=self.request.user.store
        ).select_related('category', 'store', 'created_by')

        # Add filter for low stock if requested
        if self.request.query_params.get('low_stock') == 'true':
            queryset = queryset.filter(stock__lte=F('low_stock_threshold'))

        return queryset

    def get_serializer_class(self):
        """Use list serializer for GET, full serializer for POST"""
        if self.request.method == 'GET':
            return ProductListSerializer
        return ProductSerializer

    def perform_create(self, serializer):
        """Set store and created_by automatically"""
        try:
            serializer.save(
                store=self.request.user.store,
                created_by=self.request.user
            )
            logger.info(
                f"Product '{serializer.instance.name}' created by user {self.request.user.email}"
            )
        except Exception as e:
            logger.error(f"Error creating product: {str(e)}")
            raise


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Get, update, or soft-delete a product.
    """
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated, IsSameStore]

    def get_queryset(self):
        """Return products from user's store"""
        return Product.objects.filter(
            store=self.request.user.store
        ).select_related('category', 'store', 'created_by')

    def perform_destroy(self, instance):
        """Soft delete - mark as inactive instead of deleting"""
        instance.is_active = False
        instance.save()
        logger.info(
            f"Product '{instance.name}' deactivated by user {self.request.user.email}"
        )


class LowStockProductsView(generics.ListAPIView):
    """
    Get products with stock below or equal to their low stock threshold.
    """
    serializer_class = ProductListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None  # Return all low stock items

    def get_queryset(self):
        """Return low stock products from user's store"""
        return Product.objects.filter(
            store=self.request.user.store,
            is_active=True
        ).filter(
            stock__lte=F('low_stock_threshold')
        ).select_related('category', 'store').order_by('stock')


# ============================================
# INVOICE VIEWS
# ============================================

class InvoiceListCreateView(generics.ListCreateAPIView):
    """
    List and create invoices.
    Salespeople see only their invoices, owners/managers see all.
    """
    permission_classes = [permissions.IsAuthenticated, CanCreateInvoice]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['sync_status']
    ordering_fields = ['created_at', 'total']
    ordering = ['-created_at']

    def get_serializer_class(self):
        """Use list serializer for GET, full serializer for POST"""
        if self.request.method == 'GET':
            return InvoiceListSerializer
        return InvoiceSerializer

    def get_queryset(self):
        """Return invoices based on user role"""
        user = self.request.user
        queryset = Invoice.objects.filter(store=user.store)

        # Salespeople only see their own invoices
        if user.role.name == 'salesperson':
            queryset = queryset.filter(salesperson=user)

        # Filter by date range if provided
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        if start_date:
            try:
                start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__gte=start)
            except ValueError:
                pass  # Ignore invalid date format

        if end_date:
            try:
                end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__lte=end)
            except ValueError:
                pass  # Ignore invalid date format

        # Optimize queries
        queryset = queryset.select_related(
            'store', 'salesperson', 'salesperson__role'
        ).prefetch_related('items__product')

        return queryset

    @transaction.atomic
    def perform_create(self, serializer):
        """
        Create invoice with transaction to ensure data consistency.
        """
        try:
            invoice = serializer.save(
                store=self.request.user.store,
                salesperson=self.request.user
            )
            logger.info(
                f"Invoice {invoice.invoice_number} created by user {self.request.user.email}"
            )
        except Exception as e:
            logger.error(f"Error creating invoice: {str(e)}")
            raise ValidationError({"error": "Failed to create invoice. Please try again."})


class InvoiceDetailView(generics.RetrieveAPIView):
    """
    Get invoice details with all items.
    Salespeople can only view their own invoices.
    """
    serializer_class = InvoiceSerializer
    permission_classes = [permissions.IsAuthenticated, IsSalespersonOrOwner]

    def get_queryset(self):
        """Return invoices based on user role"""
        user = self.request.user
        queryset = Invoice.objects.filter(store=user.store)

        # Salespeople only see their own invoices
        if user.role.name == 'salesperson':
            queryset = queryset.filter(salesperson=user)

        # Optimize queries
        queryset = queryset.select_related(
            'store', 'salesperson', 'salesperson__role'
        ).prefetch_related('items__product')

        return queryset


class BulkInvoiceSyncView(generics.CreateAPIView):
    """
    Sync multiple invoices from offline mode.
    Uses transactions to ensure all-or-nothing processing per invoice.
    """
    serializer_class = BulkInvoiceSyncSerializer
    permission_classes = [permissions.IsAuthenticated, CanCreateInvoice]

    def create(self, request, *args, **kwargs):
        """Process bulk invoice sync"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            # Process sync
            result = serializer.save()

            # Log sync operation
            log_status = 'completed' if result['failed'] == 0 else 'failed'
            sync_log = SyncLog.objects.create(
                user=request.user,
                store=request.user.store,
                sync_type='invoice',
                status=log_status,
                items_synced=result['synced'],
                items_failed=result['failed'],
                details=result,
                completed_at=timezone.now()
            )

            logger.info(
                f"Bulk sync completed: {result['synced']} synced, "
                f"{result['failed']} failed by user {request.user.email}"
            )

            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Bulk sync error: {str(e)}")

            # Log failed sync
            SyncLog.objects.create(
                user=request.user,
                store=request.user.store,
                sync_type='invoice',
                status='failed',
                items_synced=0,
                items_failed=0,
                error_message=str(e),
                completed_at=timezone.now()
            )

            return Response(
                {"error": "Sync failed. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ============================================
# DASHBOARD & ANALYTICS VIEWS
# ============================================

class DashboardStatsView(views.APIView):
    """
    Get dashboard statistics for store owners and managers.
    """
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    def get(self, request):
        """Calculate and return dashboard statistics"""
        try:
            store = request.user.store
            today = timezone.now().date()
            week_ago = today - timedelta(days=7)
            month_ago = today - timedelta(days=30)

            # Today's sales
            today_invoices = Invoice.objects.filter(
                store=store,
                created_at__date=today,
                sync_status='SYNCED'
            )
            today_sales = today_invoices.aggregate(
                total=Sum('total')
            )['total'] or Decimal('0.00')
            invoice_count = today_invoices.count()

            # Active salespeople today
            active_salespeople = today_invoices.values(
                'salesperson'
            ).distinct().count()

            # Week sales
            week_sales = Invoice.objects.filter(
                store=store,
                created_at__date__gte=week_ago,
                sync_status='SYNCED'
            ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')

            # Month sales
            month_sales = Invoice.objects.filter(
                store=store,
                created_at__date__gte=month_ago,
                sync_status='SYNCED'
            ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')

            # Top product today
            top_product = InvoiceItem.objects.filter(
                invoice__store=store,
                invoice__created_at__date=today,
                invoice__sync_status='SYNCED'
            ).values('product_name').annotate(
                total_quantity=Sum('quantity')
            ).order_by('-total_quantity').first()

            # Low stock products count
            low_stock_count = Product.objects.filter(
                store=store,
                is_active=True
            ).filter(stock__lte=F('low_stock_threshold')).count()

            data = {
                'today_sales': today_sales,
                'invoice_count': invoice_count,
                'top_product': top_product['product_name'] if top_product else 'N/A',
                'active_salespeople': active_salespeople,
                'week_sales': week_sales,
                'month_sales': month_sales,
                'low_stock_products': low_stock_count
            }

            serializer = DashboardStatsSerializer(data)
            return Response(serializer.data)

        except Exception as e:
            logger.error(f"Dashboard stats error: {str(e)}")
            return Response(
                {"error": "Failed to fetch dashboard statistics"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SalesReportView(views.APIView):
    """
    Get sales report by salesperson.
    Only accessible to owners and managers.
    """
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    def get(self, request):
        """Generate sales report by salesperson"""
        try:
            store = request.user.store

            # Get date range from query params
            start_date = request.query_params.get(
                'start_date',
                (timezone.now().date() - timedelta(days=30)).isoformat()
            )
            end_date = request.query_params.get(
                'end_date',
                timezone.now().date().isoformat()
            )

            # Sales by salesperson
            sales_data = Invoice.objects.filter(
                store=store,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date,
                sync_status='SYNCED'
            ).values(
                'salesperson__id',
                'salesperson__name'
            ).annotate(
                total_sales=Sum('total'),
                invoice_count=Count('id'),
                average_sale=Avg('total')
            ).order_by('-total_sales')

            # Format data
            report_data = [
                {
                    'salesperson_id': item['salesperson__id'],
                    'salesperson_name': item['salesperson__name'],
                    'total_sales': item['total_sales'],
                    'invoice_count': item['invoice_count'],
                    'average_sale': item['average_sale']
                }
                for item in sales_data
            ]

            serializer = SalesReportSerializer(report_data, many=True)
            return Response(serializer.data)

        except Exception as e:
            logger.error(f"Sales report error: {str(e)}")
            return Response(
                {"error": "Failed to generate sales report"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProductReportView(views.APIView):
    """
    Get product sales report.
    Only accessible to owners and managers.
    """
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    def get(self, request):
        """Generate product sales report"""
        try:
            store = request.user.store

            # Get date range and limit from query params
            start_date = request.query_params.get(
                'start_date',
                (timezone.now().date() - timedelta(days=30)).isoformat()
            )
            end_date = request.query_params.get(
                'end_date',
                timezone.now().date().isoformat()
            )
            limit = int(request.query_params.get('limit', 20))

            # Product sales data
            product_data = InvoiceItem.objects.filter(
                invoice__store=store,
                invoice__created_at__date__gte=start_date,
                invoice__created_at__date__lte=end_date,
                invoice__sync_status='SYNCED'
            ).values(
                'product__id',
                'product__name',
                'product__code'
            ).annotate(
                quantity_sold=Sum('quantity'),
                total_revenue=Sum('total')
            ).order_by('-quantity_sold')[:limit]

            # Format data
            report_data = [
                {
                    'product_id': item['product__id'],
                    'product_name': item['product__name'],
                    'product_code': item['product__code'],
                    'quantity_sold': item['quantity_sold'],
                    'total_revenue': item['total_revenue']
                }
                for item in product_data
            ]

            serializer = ProductReportSerializer(report_data, many=True)
            return Response(serializer.data)

        except Exception as e:
            logger.error(f"Product report error: {str(e)}")
            return Response(
                {"error": "Failed to generate product report"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ============================================
# SYNC VIEWS
# ============================================

class SyncStatusView(views.APIView):
    """
    Get sync status for authenticated user.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Return current sync status"""
        try:
            user = request.user

            # Count pending invoices
            pending_invoices = Invoice.objects.filter(
                salesperson=user,
                sync_status='PENDING'
            ).count()

            # Get last successful sync
            last_sync = SyncLog.objects.filter(
                user=user,
                status='completed'
            ).order_by('-completed_at').first()

            data = {
                'pending_invoices': pending_invoices,
                'last_sync_time': last_sync.completed_at if last_sync else None,
                'sync_status': 'online' if pending_invoices == 0 else 'pending'
            }

            return Response(data)

        except Exception as e:
            logger.error(f"Sync status error: {str(e)}")
            return Response(
                {"error": "Failed to fetch sync status"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SyncHistoryView(generics.ListAPIView):
    """
    Get sync history for authenticated user.
    """
    serializer_class = SyncLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None  # Return all recent logs

    def get_queryset(self):
        """Return last 20 sync logs for user"""
        return SyncLog.objects.filter(
            user=self.request.user
        ).select_related('store').order_by('-started_at')[:20]


# ============================================
# PROFILE VIEWS
# ============================================

class UserProfileView(generics.RetrieveUpdateAPIView):
    """
    Get and update user profile.
    Users can only update certain fields.
    """
    serializer_class = UserProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        """Return authenticated user"""
        return self.request.user


@staff_member_required
def download_product_template(request):
    """
    Download Excel template for bulk product upload with instructions and sample data.
    """
    # Check if user has a store (skip for superusers)
    if not request.user.is_superuser and not request.user.store:
        messages.error(request, 'You must be assigned to a store to download the template.')
        return redirect('admin:pos_app_product_changelist')

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ===== INSTRUCTIONS SHEET =====
    ws_instructions = wb.create_sheet('Instructions', 0)

    # Header
    ws_instructions['A1'] = 'Bulk Product Upload Instructions'
    ws_instructions['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws_instructions['A1'].fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    ws_instructions.merge_cells('A1:D1')
    ws_instructions.row_dimensions[1].height = 30
    ws_instructions['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Instructions content
    instructions = [
        ('', ''),
        ('STEP-BY-STEP GUIDE', ''),
        ('1. Fill Product Data', 'Go to the "Products" sheet and fill in your product information'),
        ('2. Follow Format', 'Each column has specific requirements (see below)'),
        ('3. Check Categories', 'Use category names from the "Categories" sheet'),
        ('4. Save File', 'Save the file as .xlsx format'),
        ('5. Upload', 'Go back to admin and upload the file'),
        ('', ''),
        ('COLUMN DESCRIPTIONS', ''),
        ('code *', 'Unique product code/SKU (REQUIRED) - e.g., PROD001'),
        ('name *', 'Product name (REQUIRED) - e.g., Laptop Dell XPS 15'),
        ('description', 'Product description (OPTIONAL)'),
        ('category_name', 'Category name - Must match existing category (see Categories sheet)'),
        ('price *', 'Selling price (REQUIRED) - Must be positive number - e.g., 1500000'),
        ('cost', 'Cost price (OPTIONAL) - Must be positive number - e.g., 1200000'),
        ('stock *', 'Initial stock quantity (REQUIRED) - e.g., 50'),
        ('low_stock_threshold', 'Alert when stock falls below this (OPTIONAL, default: 10)'),
        ('barcode', 'Product barcode (OPTIONAL) - Must be unique if provided'),
        ('image_url', 'Product image URL (OPTIONAL)'),
        ('', ''),
        ('IMPORTANT NOTES', ''),
        ('✓', 'Fields marked with * are REQUIRED'),
        ('✓', 'Code and barcode must be UNIQUE across all products'),
        ('✓', 'Prices must be positive numbers (decimals allowed)'),
        ('✓', 'Stock must be a positive whole number'),
        ('✓', 'Category name must match an existing category exactly'),
        ('✓', 'Delete sample data before adding your products'),
        ('✓', 'Do not modify column headers in the Products sheet'),
    ]

    for idx, (col1, col2) in enumerate(instructions, start=2):
        ws_instructions[f'A{idx}'] = col1
        ws_instructions[f'B{idx}'] = col2

        # Style for headers
        if col1 in ['STEP-BY-STEP GUIDE', 'COLUMN DESCRIPTIONS', 'IMPORTANT NOTES']:
            ws_instructions[f'A{idx}'].font = Font(bold=True, size=12, color='1F2937')
            ws_instructions[f'A{idx}'].fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            ws_instructions.merge_cells(f'A{idx}:B{idx}')

    # Set column widths
    ws_instructions.column_dimensions['A'].width = 25
    ws_instructions.column_dimensions['B'].width = 70

    # ===== PRODUCTS SHEET =====
    ws_products = wb.create_sheet('Products', 1)

    # Headers
    headers = [
        'code', 'name', 'description', 'category_name',
        'price', 'cost', 'stock', 'low_stock_threshold',
        'barcode', 'image_url'
    ]

    # Style headers
    header_fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws_products.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Sample data
    sample_products = [
        ['LAPTOP001', 'Dell XPS 15', 'High-performance laptop with 16GB RAM', 'Electronics',
         '1500000', '1200000', '25', '5', '123456789001', 'https://example.com/laptop.jpg'],
        ['PHONE001', 'Samsung Galaxy S23', 'Latest smartphone with 5G', 'Electronics',
         '800000', '650000', '50', '10', '123456789002', 'https://example.com/phone.jpg'],
        ['DESK001', 'Office Desk Large', 'Wooden office desk 160x80cm', 'Furniture',
         '350000', '250000', '15', '3', '123456789003', 'https://example.com/desk.jpg'],
    ]

    for row_num, product in enumerate(sample_products, 2):
        for col_num, value in enumerate(product, 1):
            cell = ws_products.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    # Set column widths
    column_widths = {
        'A': 15, 'B': 30, 'C': 40, 'D': 20,
        'E': 12, 'F': 12, 'G': 10, 'H': 20,
        'I': 18, 'J': 35
    }

    for col, width in column_widths.items():
        ws_products.column_dimensions[col].width = width

    # Freeze header row
    ws_products.freeze_panes = 'A2'

    # ===== CATEGORIES SHEET =====
    ws_categories = wb.create_sheet('Categories', 2)

    # Header
    ws_categories['A1'] = 'Available Categories'
    ws_categories['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws_categories['A1'].fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    ws_categories['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_categories.row_dimensions[1].height = 25

    ws_categories['B1'] = 'Store'
    ws_categories['B1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws_categories['B1'].fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    ws_categories['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # Get categories from database
    from .models import Category

    # For superusers, show all categories; for others, only their store
    if request.user.is_superuser:
        categories = Category.objects.filter(is_active=True).select_related('store').order_by('store__name', 'name')
    else:
        categories = Category.objects.filter(
            is_active=True,
            store=request.user.store
        ).select_related('store').order_by('name')

    for idx, category in enumerate(categories, 2):
        ws_categories[f'A{idx}'] = category.name
        ws_categories[f'B{idx}'] = category.store.name if category.store else 'N/A'

        ws_categories[f'A{idx}'].alignment = Alignment(horizontal='left', vertical='center')
        ws_categories[f'B{idx}'].alignment = Alignment(horizontal='left', vertical='center')

    ws_categories.column_dimensions['A'].width = 30
    ws_categories.column_dimensions['B'].width = 25

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=product_upload_template.xlsx'

    wb.save(response)
    return response


@staff_member_required
def bulk_upload_products(request):
    """
    Handle bulk product upload from Excel file.
    """
    from .models import Product, Category

    # Check if user has a store (skip for superusers)
    if not request.user.is_superuser and not request.user.store:
        messages.error(request, 'You must be assigned to a store to upload products.')
        return redirect('admin:pos_app_product_changelist')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, 'Please select an Excel file to upload.')
            return redirect('admin:pos_app_product_bulk_upload')

        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload .xlsx or .xls file.')
            return redirect('admin:pos_app_product_bulk_upload')

        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file)

            # Get Products sheet
            if 'Products' not in wb.sheetnames:
                messages.error(request, 'Invalid template. "Products" sheet not found.')
                return redirect('admin:pos_app_product_bulk_upload')

            ws = wb['Products']

            # Get headers
            headers = [cell.value for cell in ws[1]]

            # Validate headers
            required_headers = ['code', 'name', 'price', 'stock']
            for header in required_headers:
                if header not in headers:
                    messages.error(request, f'Missing required column: {header}')
                    return redirect('admin:pos_app_product_bulk_upload')

            # Get user's store
            user_store = request.user.store

            # Process products
            products_to_create = []
            errors = []
            success_count = 0
            skip_count = 0

            with transaction.atomic():
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip empty rows
                    if not any(row):
                        skip_count += 1
                        continue

                    # Create dict from row
                    row_data = dict(zip(headers, row))

                    # Validate required fields
                    if not row_data.get('code'):
                        errors.append(f"Row {idx}: Missing product code")
                        continue

                    if not row_data.get('name'):
                        errors.append(f"Row {idx}: Missing product name")
                        continue

                    if not row_data.get('price'):
                        errors.append(f"Row {idx}: Missing price")
                        continue

                    # Check for duplicate code
                    if Product.objects.filter(code=row_data['code'], store=user_store).exists():
                        errors.append(f"Row {idx}: Product code '{row_data['code']}' already exists")
                        continue

                    # Check for duplicate barcode
                    if row_data.get('barcode'):
                        if Product.objects.filter(barcode=row_data['barcode'], store=user_store).exists():
                            errors.append(f"Row {idx}: Barcode '{row_data['barcode']}' already exists")
                            continue

                    # Get or validate category
                    category = None
                    if row_data.get('category_name'):
                        try:
                            category = Category.objects.get(
                                name=row_data['category_name'],
                                store=user_store,
                                is_active=True
                            )
                        except Category.DoesNotExist:
                            errors.append(
                                f"Row {idx}: Category '{row_data['category_name']}' not found. "
                                f"Please use exact category names from the Categories sheet."
                            )
                            continue

                    # Create product
                    try:
                        product = Product(
                            store=user_store,
                            code=row_data['code'],
                            name=row_data['name'],
                            description=row_data.get('description', ''),
                            category=category,
                            price=Decimal(str(row_data['price'])),
                            cost=Decimal(str(row_data['cost'])) if row_data.get('cost') else None,
                            stock=int(row_data.get('stock', 0)),
                            low_stock_threshold=int(row_data.get('low_stock_threshold', 10)),
                            barcode=row_data.get('barcode', ''),
                            image_url=row_data.get('image_url', ''),
                            created_by=request.user,
                            is_active=True
                        )

                        product.full_clean()  # Validate
                        products_to_create.append(product)
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                        continue

                # Bulk create products
                if products_to_create:
                    Product.objects.bulk_create(products_to_create)

            # Show results
            if success_count > 0:
                messages.success(
                    request,
                    f'✓ Successfully uploaded {success_count} product(s)!'
                )

            if skip_count > 0:
                messages.info(request, f'Skipped {skip_count} empty row(s)')

            if errors:
                error_message = f'Failed to upload {len(errors)} product(s):\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_message += f'\n... and {len(errors) - 10} more errors'
                messages.warning(request, error_message)

            return redirect('admin:pos_app_product_changelist')

        except Exception as e:
            logger.error(f"Bulk upload error: {str(e)}", exc_info=True)
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('admin:pos_app_product_bulk_upload')

    # GET request - show upload form
    return render(request, 'admin/pos_app/product/bulk_upload.html')


@staff_member_required
def export_products_excel(request):
    """
    Export all products to Excel file.
    """
    from .models import Product

    # Check if user has a store (skip for superusers)
    if not request.user.is_superuser and not request.user.store:
        messages.error(request, 'You must be assigned to a store to export products.')
        return redirect('admin:pos_app_product_changelist')

    # Get user's store products
    if request.user.is_superuser:
        products = Product.objects.all().select_related('category', 'store').order_by('store__name', 'code')
    else:
        products = Product.objects.filter(
            store=request.user.store
        ).select_related('category', 'store').order_by('code')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'

    # Headers
    headers = [
        'Code', 'Name', 'Description', 'Category',
        'Price', 'Cost', 'Stock', 'Low Stock Threshold',
        'Barcode', 'Image URL', 'Is Active', 'Created At'
    ]

    # Style headers
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.code)
        ws.cell(row=row_num, column=2, value=product.name)
        ws.cell(row=row_num, column=3, value=product.description)
        ws.cell(row=row_num, column=4, value=product.category.name if product.category else '')
        ws.cell(row=row_num, column=5, value=float(product.price))
        ws.cell(row=row_num, column=6, value=float(product.cost) if product.cost else '')
        ws.cell(row=row_num, column=7, value=product.stock)
        ws.cell(row=row_num, column=8, value=product.low_stock_threshold)
        ws.cell(row=row_num, column=9, value=product.barcode)
        ws.cell(row=row_num, column=10, value=product.image_url)
        ws.cell(row=row_num, column=11, value='Yes' if product.is_active else 'No')
        ws.cell(row=row_num, column=12, value=product.created_at.strftime('%Y-%m-%d %H:%M:%S'))

    # Set column widths
    column_widths = [15, 30, 40, 20, 12, 12, 10, 20, 18, 35, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Safe filename
    if request.user.is_superuser:
        filename = 'products_export_all.xlsx'
    else:
        store_code = request.user.store.code if request.user.store else 'unknown'
        filename = f'products_export_{store_code}.xlsx'

    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

# ============================================
# UTILITY VIEWS
# ============================================

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def health_check(request):
    """
    Health check endpoint to verify API and user authentication.
    """
    return Response({
        'status': 'healthy',
        'timestamp': timezone.now().isoformat(),
        'user': {
            'email': request.user.email,
            'name': request.user.name,
            'store': request.user.store.name if request.user.store else None,
            'role': request.user.role.name if request.user.role else None
        }
    })